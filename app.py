# app.py
# Refactored Flask application for concurrent, time-limited data processing.
# Updated to use gemini-2.5-flash for high-speed, cost-effective analysis.

import os
import io
import json
import re
import concurrent.futures
from flask import Flask, request, jsonify
from dotenv import load_dotenv
import requests
import traceback
import base64
import csv
import pandas as pd
import xml.etree.ElementTree as ET
import google.generativeai as genai
from bs4 import BeautifulSoup
import openpyxl
import PyPDF2
import pandas as pd
import yaml
from google.api_core.exceptions import ResourceExhausted
from PIL import Image
from dotenv import load_dotenv

# Load environment variables from a .env file
load_dotenv()

# --- API Key Configuration ---
# NOTE: The original code used GEMINI_API_KEY. If you were to use a different
# service that requires an AIPIPE_TOKEN, you would configure it here.
# For this Gemini-based script, we continue to use the GEMINI_API_KEY.
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")

# Configure the Gemini API with the loaded key
genai.configure(api_key=GEMINI_API_KEY)

# --- Updated Model Initialization ---
def get_gemini_model():
    """Initializes and returns the gemini-2.5-flash model for high-speed tasks."""
    try:
        model = genai.GenerativeModel('gemini-2.5-flash')
        # A quick test to ensure the model is accessible
        model.generate_content("test")
        print("Using gemini-2.5-flash model for high-speed tasks.")
        return model
    except Exception as e:
        # If model initialization or a test call fails, log the error.
        print(f"An error occurred while initializing the model: {e}")
        return None

# --- File Handlers (with new image handler) ---
# These functions are unchanged from your original script as they are robust.
def handle_text_file(file_content): return file_content.decode('utf-8')
def handle_html_file(file_content): return BeautifulSoup(file_content, 'html.parser').get_text(separator=' ', strip=True)
def handle_markdown_file(file_content): return file_content.decode('utf-8')
def handle_csv_file(file_content):
    csv_reader = csv.reader(io.StringIO(file_content.decode('utf-8')))
    try:
        header = next(csv_reader)
        rows = [",".join(row) for row in csv_reader]
        return ",".join(header) + "\n" + "\n".join(rows)
    except StopIteration:
        return "CSV file is empty."
def handle_json_file(file_content): return json.dumps(json.loads(file_content), indent=2)
def handle_pdf_file(file_content):
    pdf_stream = io.BytesIO(file_content)
    text = ""
    try:
        reader = PyPDF2.PdfReader(pdf_stream)
        for page in reader.pages:
            text += page.extract_text() or ""
    except Exception as e:
        return f"Error processing PDF: {str(e)}"
    return text
def handle_excel_file(file_content):
    stream = io.BytesIO(file_content)
    data = ""
    try:
        workbook = openpyxl.load_workbook(stream, read_only=True)
        for sheet in workbook.sheetnames:
            sh = workbook[sheet]
            data += f"\n--- Sheet: {sheet} ---\n"
            for row in sh.iter_rows():
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                data += ",".join(row_data) + "\n"
    except Exception as e:
        return f"Error processing Excel: {str(e)}"
    return data
def handle_sql_file(file_content): return file_content.decode('utf-8')
def handle_xml_file(file_content):
    try:
        root = ET.fromstring(file_content)
        return ET.tostring(root, encoding='unicode', method='xml')
    except ET.ParseError as e:
        return f"Error parsing XML: {str(e)}"
def handle_parquet_file(file_content):
    try:
        df = pd.read_parquet(io.BytesIO(file_content))
        return df.to_json(orient="records", indent=2)
    except Exception as e:
        return f"Error processing Parquet: {str(e)}"
def handle_yaml_file(file_content):
    try:
        data = yaml.safe_load(file_content)
        return yaml.dump(data, sort_keys=False)
    except yaml.YAMLError as e:
        return f"Error processing YAML: {str(e)}"
def handle_image_file(file_content, filename):
    try:
        if file_content.startswith(b'data:image'):
            base64_data = file_content.split(b',')[1].decode('utf-8')
            mime_type = file_content.split(b';')[0].split(b':')[1].decode('utf-8')
            return ('image_url', {"mime_type": mime_type, "data": base64_data}, filename)
        
        img_bytes = io.BytesIO(file_content)
        img_bytes.seek(0)
        base64_data = base64.b64encode(img_bytes.read()).decode('utf-8')
        mime_type = "image/png"
        return ('image_url', {"mime_type": mime_type, "data": base64_data}, filename)
        
    except Exception as e:
        return ('text', f"Error processing image with Pillow or Base64 encoding: {str(e)}", filename)

# Mapping of file extensions to their respective handlers
FILE_HANDLERS = {
    '.txt': handle_text_file, '.html': handle_html_file, '.md': handle_markdown_file,
    '.csv': handle_csv_file, '.json': handle_json_file, '.pdf': handle_pdf_file,
    '.xlsx': handle_excel_file, '.xls': handle_excel_file, '.sql': handle_sql_file,
    '.xml': handle_xml_file, '.parquet': handle_parquet_file, '.yaml': handle_yaml_file,
    '.png': handle_image_file, '.jpg': handle_image_file, '.jpeg': handle_image_file, '.webp': handle_image_file
}

IMAGE_EXTENSIONS = ['.png', '.jpg', '.jpeg', '.webp']
# --- Core Logic Functions (Refactored for Multimodality) ---
def analyze_with_gemini(questions, text_context, image_parts, gemini_model):
    if not gemini_model:
        return json.dumps({"error": "Gemini Model not initialized"})
    
    # This prompt is the instruction for the AI. It's crucial for getting a good response.
    full_text_prompt =  f""" You are an intelligent data analyst. Your ONLY task is to return a valid JSON response that answers the questions using the provided data (which may include text and images). 
STRICT RULES:
- Answer strictly as a JSON. 
- If answer is number, then return it as number.dont wrab number with quotes.
- Strings must always be quoted and properly escaped. 
- Visualizations must be returned as 100kb base64-encoded strings inside the JSON. Always avoid untermination.
- Do NOT include any explanations, comments, text, or code before or after the JSON.
- Do NOT use markdown formatting or code fences.
- You should return your response which is in correct json syntax.

--- Questions to Answer ---
{questions}

--- Data Context ---
{text_context}

Return ONLY a valid JSON.
"""
    prompt_parts = [full_text_prompt] + image_parts
    try:
        response = gemini_model.generate_content(
            prompt_parts,
            generation_config={"response_mime_type": "application/json"}
        )
        return response.text
    except Exception as e:
        return json.dumps({"error": f"Gemini API Error: {str(e)}"})
def robust_json_parser(text_response: str):
    """
    Cleans LLM output and extracts JSON safely.
    Ensures final result is always valid JSON (object or array).
    """
    import json, re

    if not isinstance(text_response, str):
        text_response = str(text_response)

    # Step 1: Strip markdown fences and leading/trailing spaces
    cleaned = re.sub(r"^```(?:json)?|```$", "", text_response.strip(),
                     flags=re.MULTILINE | re.DOTALL).strip()
     # Step 4: Balance brackets (in case output is cut off)
    if cleaned.startswith("[") and not cleaned.endswith("]"):
        cleaned += '"]'
    elif cleaned.startswith("{") and not cleaned.endswith("}"):
        cleaned +='"}'

    # Step 2: Try to locate the first JSON object/array inside the text
    match = re.search(r'(\{.*\}|\[.*\])', cleaned, re.DOTALL)
    if match:
        cleaned = match.group(1)

    # Step 3: Cleanup common issues
    cleaned = re.sub(r",\s*([\]}])", r"\1", cleaned)   # remove trailing commas
    cleaned = cleaned.replace("\n", "").replace("\r", "")  # flatten newlines
    cleaned = cleaned.replace("'", '"')  # replace single quotes with double
    
    # Step 5: Attempt JSON parsing
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError as e:
        # Final fallback: try to recover with regex fixes
        try:
            fixed = re.sub(r'\\(?!["\\/bfnrt])', r"\\\\", cleaned)  # escape bad backslashes
            return json.loads(fixed)
        except Exception:
            return {
                "error": "Could not parse JSON",
                "details": str(e),
                "raw_snippet": cleaned  # show first 200 chars for debugging
            }


def process_single_file(file_storage):
    """Processes a single file based on its extension."""
    filename = file_storage.filename
    file_extension = os.path.splitext(filename)[1].lower()
    
    if file_extension in FILE_HANDLERS:
        try:
            file_storage.seek(0)
            file_content = file_storage.read()
            handler = FILE_HANDLERS[file_extension]
            if file_extension in IMAGE_EXTENSIONS:
                return handler(file_content, filename)
            else:
                processed_data = handler(file_content)
                text_data = f"\n--- Content from {filename} ---\n{processed_data}\n"
                return ('text', text_data, filename)
        except Exception as e:
            error_text = f"\n--- Error processing {filename}: {str(e)} ---\n"
            return ('text', error_text, filename)
            
    return ('text', '', filename)

def process_request_worker(request_files, model_instance):
    """A worker function to process the entire request."""
    questions_file_content = request_files['questions.txt'].read().decode('utf-8')
    questions = questions_file_content
    
    text_context_parts = []
    image_parts = []
    
    data_files = [f for k, f in request_files.items() if k != 'questions.txt']
    
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_file = {executor.submit(process_single_file, f): f for f in data_files}
        for future in concurrent.futures.as_completed(future_to_file):
            try:
                content_type, data, filename = future.result()
                if content_type == 'image_url':
                    image_parts.append(genai.types.Part.from_data(
                        data=data["data"], 
                        mime_type=data["mime_type"]
                    ))
                else:
                    text_context_parts.append(data)
            except Exception as e:
                filename = future_to_file[future].filename
                print(f"Error processing file {filename} in thread: {e}")
                text_context_parts.append(f"\n--- Exception processing {filename}: {e} ---\n")

    final_text_context = "".join(text_context_parts)
    
    analysis_result_text = analyze_with_gemini(questions, final_text_context, image_parts, model_instance)
    
    final_json_object = robust_json_parser(analysis_result_text)
    
    return final_json_object

# --- API Endpoint (Controller - Unchanged) ---
from flask import Flask, request, jsonify
import concurrent.futures
import os

app = Flask(__name__)

@app.route('/api/', methods=['POST'])
def data_analyst_agent():
    if 'questions.txt' not in request.files:
        return jsonify({"error": "questions.txt file is required."}), 400

    model_instance = get_gemini_model()

    # No timeout anymore
    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(process_request_worker, request.files, model_instance)
        try:
            result = future.result() 
            if isinstance(result, list):
                print("✅ Result is a JSON array")
            elif isinstance(result, dict):
                print("✅ Result is a JSON object")
            else:
                print("⚠️ Result is neither array nor object:", type(result))
            return jsonify(result), 200
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            return jsonify({"error":str(e)})


# ------------------ Health Check ------------------
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({"status": "ok"}), 200


# ------------------ Main (for running with a production server) ------------------
if __name__ == '__main__':
    from waitress import serve
    port = int(os.environ.get('PORT', 5000))
    print(f"Starting server on http://0.0.0.0:{port}")
    serve(app, host='0.0.0.0', port=port, threads=8)
